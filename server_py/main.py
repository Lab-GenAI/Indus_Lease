import os
import sys
import json
import asyncio
import time
import io
import threading
import tempfile
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional, List
from dotenv import load_dotenv

load_dotenv()

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request, BackgroundTasks
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

from server_py.db import execute_query, execute_no_fetch
from server_py.seed import seed_database
from server_py import storage
from server_py.document_parser import get_file_extension, is_supported_file, parse_document, render_email_as_html
from server_py.extractor import extract_tags_for_lease as extract_tags_from_lease
from server_py.progress import emit_progress, subscribe, unsubscribe, generate_task_id

UPLOAD_DIR = os.path.join(tempfile.gettempdir(), "lease_uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

EXTRACTION_SEMAPHORE = threading.Semaphore(5)
EXTRACTION_TIMEOUT_MINUTES = 60


def _is_extraction_stale(extraction) -> bool:
    if not extraction or extraction["status"] != "processing":
        return False
    ts = extraction.get("updated_at") or extraction.get("created_at")
    if not ts:
        return False
    from datetime import timezone
    try:
        if isinstance(ts, str):
            ts_dt = datetime.fromisoformat(ts)
        else:
            ts_dt = ts
        if ts_dt.tzinfo is None:
            ts_dt = ts_dt.replace(tzinfo=timezone.utc)
        age_minutes = (datetime.now(timezone.utc) - ts_dt).total_seconds() / 60
        return age_minutes >= EXTRACTION_TIMEOUT_MINUTES
    except Exception:
        return False

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def log(message: str, source: str = "fastapi"):
    now = datetime.now().strftime("%I:%M:%S %p")
    print(f"{now} [{source}] {message}")


def _cleanup_stale_temp_dirs():
    import glob
    tmp = tempfile.gettempdir()
    patterns = ["vision-*", "ocr-*", "email-att-*"]
    cleaned = 0
    for pattern in patterns:
        for d in glob.glob(os.path.join(tmp, pattern)):
            try:
                if os.path.isdir(d):
                    age_hours = (time.time() - os.path.getmtime(d)) / 3600
                    if age_hours > 1:
                        shutil.rmtree(d, ignore_errors=True)
                        cleaned += 1
            except Exception:
                pass
    if cleaned:
        print(f"[CLEANUP] Removed {cleaned} stale temp directories")


@app.on_event("startup")
async def startup_event():
    _cleanup_stale_temp_dirs()
    seed_database()
    try:
        execute_no_fetch("""
            ALTER TABLE extractions ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        """)
    except Exception:
        pass
    try:
        execute_no_fetch("""
            CREATE TABLE IF NOT EXISTS app_settings (
                id SERIAL PRIMARY KEY,
                key TEXT NOT NULL UNIQUE,
                value TEXT NOT NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP NOT NULL
            )
        """)
    except Exception:
        pass
    try:
        execute_no_fetch("""
            CREATE TABLE IF NOT EXISTS cost_logs (
                id SERIAL PRIMARY KEY,
                type TEXT NOT NULL,
                lease_id INTEGER REFERENCES leases(id) ON DELETE CASCADE,
                site_id INTEGER REFERENCES sites(id) ON DELETE CASCADE,
                model TEXT NOT NULL,
                input_tokens INTEGER NOT NULL DEFAULT 0,
                output_tokens INTEGER NOT NULL DEFAULT 0,
                total_tokens INTEGER NOT NULL DEFAULT 0,
                cost_usd REAL NOT NULL DEFAULT 0,
                cost_inr REAL NOT NULL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP NOT NULL
            )
        """)
    except Exception:
        pass
    try:
        execute_no_fetch("CREATE INDEX IF NOT EXISTS idx_leases_site_id ON leases(site_id)")
        execute_no_fetch("CREATE INDEX IF NOT EXISTS idx_leases_site_lease ON leases(site_id, lease_number)")
        execute_no_fetch("CREATE INDEX IF NOT EXISTS idx_files_lease_id ON files(lease_id)")
        execute_no_fetch("CREATE INDEX IF NOT EXISTS idx_files_dedup ON files(lease_id, file_name, file_size)")
        execute_no_fetch("CREATE INDEX IF NOT EXISTS idx_extractions_lease_id ON extractions(lease_id)")
        execute_no_fetch("CREATE INDEX IF NOT EXISTS idx_cost_logs_site_id ON cost_logs(site_id)")
        execute_no_fetch("CREATE INDEX IF NOT EXISTS idx_cost_logs_lease_id ON cost_logs(lease_id)")
    except Exception:
        pass
    log("Server started")


@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = time.time()
    response = await call_next(request)
    if request.url.path.startswith("/api"):
        duration = int((time.time() - start) * 1000)
        log(f"{request.method} {request.url.path} {response.status_code} in {duration}ms")
    if response.status_code == 204:
        from starlette.responses import Response as StarletteResponse
        return StarletteResponse(status_code=204)
    return response


@app.get("/api/dashboard/stats")
async def dashboard_stats():
    try:
        return storage.get_dashboard_stats()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/sites")
async def list_sites():
    try:
        return storage.get_sites()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/sites/{site_id}")
async def get_site(site_id: int):
    try:
        site = storage.get_site(site_id)
        if not site:
            raise HTTPException(status_code=404, detail="Site not found")

        leases_list = storage.get_leases_by_site(site_id)
        leases_with_details = []
        for lease in leases_list:
            lease_files = storage.get_files_by_lease(lease["id"])
            extraction = storage.get_extraction_by_lease(lease["id"])
            lease_data = {
                "id": lease["id"],
                "siteId": lease["site_id"],
                "leaseNumber": lease["lease_number"],
                "status": lease["status"],
                "createdAt": lease["created_at"].isoformat() if lease.get("created_at") else None,
                "files": [{
                    "id": f["id"],
                    "leaseId": f["lease_id"],
                    "fileName": f["file_name"],
                    "fileType": f["file_type"],
                    "filePath": f["file_path"],
                    "fileSize": f["file_size"],
                    "createdAt": f["created_at"].isoformat() if f.get("created_at") else None,
                } for f in lease_files],
                "extraction": _map_extraction_response(extraction) if extraction else None,
            }
            leases_with_details.append(lease_data)

        return {
            "id": site["id"],
            "siteId": site["site_id"],
            "createdAt": site["created_at"].isoformat() if site.get("created_at") else None,
            "leases": leases_with_details,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/sites")
async def delete_all_sites():
    try:
        storage.delete_all_sites()
        return JSONResponse(status_code=204, content=None)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/sites/{site_id}")
async def delete_site(site_id: int):
    try:
        site = storage.get_site(site_id)
        if not site:
            raise HTTPException(status_code=404, detail="Site not found")
        storage.delete_site(site_id)
        return JSONResponse(status_code=204, content=None)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/upload-folder")
async def upload_folder(
    files: List[UploadFile] = File(...),
    paths: List[str] = Form(...),
):
    try:
        if not files:
            raise HTTPException(status_code=400, detail="No files uploaded")

        sites_created = 0
        leases_created = 0
        files_created = 0
        skipped_duplicates = 0

        site_cache = {}
        lease_cache = {}
        existing_files_cache = {}

        def get_or_create_site(name):
            nonlocal sites_created
            if name in site_cache:
                return site_cache[name]
            site = storage.get_site_by_name(name)
            if not site:
                site = storage.create_site(name)
                sites_created += 1
            site_cache[name] = site
            return site

        def get_or_create_lease(site_id, lease_num):
            nonlocal leases_created
            cache_key = (site_id, lease_num)
            if cache_key in lease_cache:
                return lease_cache[cache_key]
            lease = storage.get_lease_by_number(site_id, lease_num)
            if not lease:
                lease = storage.create_lease(site_id, lease_num)
                leases_created += 1
            lease_cache[cache_key] = lease
            return lease

        def check_duplicate(lease_id, file_name, file_size):
            if lease_id not in existing_files_cache:
                existing = storage.get_files_by_lease(lease_id)
                existing_files_cache[lease_id] = {
                    (f["file_name"], f["file_size"]) for f in existing
                }
            return (file_name, file_size) in existing_files_cache[lease_id]

        os.makedirs(UPLOAD_DIR, exist_ok=True)

        bulk_rows = []

        for i, upload_file in enumerate(files):
            relative_path = paths[i] if i < len(paths) else upload_file.filename

            parts = [p for p in relative_path.split("/") if p]
            if len(parts) < 2:
                continue

            site_id_name = parts[0]
            lease_number = parts[1]
            file_name = parts[-1]

            if not is_supported_file(file_name):
                continue

            site = get_or_create_site(site_id_name)
            lease = get_or_create_lease(site["id"], lease_number)

            content = await upload_file.read()

            if check_duplicate(lease["id"], file_name, len(content)):
                skipped_duplicates += 1
                continue

            unique_suffix = f"{int(time.time() * 1000)}-{i}"
            safe_name = os.path.basename(file_name).replace(" ", "_")
            saved_filename = f"{unique_suffix}-{safe_name}"
            saved_path = os.path.join(UPLOAD_DIR, saved_filename)

            with open(saved_path, "wb") as f:
                f.write(content)

            file_type = get_file_extension(file_name)
            bulk_rows.append((lease["id"], file_name, file_type, saved_path, len(content)))
            existing_files_cache.setdefault(lease["id"], set()).add((file_name, len(content)))
            files_created += 1

        if bulk_rows:
            storage.create_files_bulk(bulk_rows)

        task_id = generate_task_id()

        dup_msg = f", {skipped_duplicates} duplicates skipped" if skipped_duplicates else ""
        emit_progress({
            "taskId": task_id, "type": "upload", "status": "completed",
            "current": files_created, "total": files_created,
            "message": f"Upload complete: {files_created} files stored{dup_msg}"
        })

        return {
            "message": "Folder processed successfully",
            "sitesCreated": sites_created,
            "leasesCreated": leases_created,
            "filesCreated": files_created,
            "skippedDuplicates": skipped_duplicates,
            "taskId": task_id,
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/tags")
async def list_tags():
    try:
        tags_list = storage.get_tags()
        return [{
            "id": t["id"],
            "name": t["name"],
            "description": t["description"],
            "category": t["category"],
            "createdAt": t["created_at"].isoformat() if t.get("created_at") else None,
        } for t in tags_list]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tags")
async def create_tag(request: Request):
    try:
        body = await request.json()
        name = body.get("name")
        if not name:
            raise HTTPException(status_code=400, detail="Tag name is required")
        tag = storage.create_tag(name, body.get("description"), body.get("category"))
        return JSONResponse(status_code=201, content={
            "id": tag["id"],
            "name": tag["name"],
            "description": tag["description"],
            "category": tag["category"],
            "createdAt": tag["created_at"].isoformat() if tag.get("created_at") else None,
        })
    except HTTPException:
        raise
    except Exception as e:
        if "unique" in str(e).lower() or "23505" in str(e):
            raise HTTPException(status_code=409, detail="A tag with this name already exists")
        raise HTTPException(status_code=500, detail=str(e))


@app.patch("/api/tags/{tag_id}")
async def update_tag(tag_id: int, request: Request):
    try:
        body = await request.json()
        tag = storage.update_tag(tag_id, body.get("name"), body.get("description"), body.get("category"))
        return {
            "id": tag["id"],
            "name": tag["name"],
            "description": tag["description"],
            "category": tag["category"],
            "createdAt": tag["created_at"].isoformat() if tag.get("created_at") else None,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/tags")
async def delete_all_tags():
    try:
        storage.delete_all_tags()
        return JSONResponse(status_code=204, content=None)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/tags/{tag_id}")
async def delete_tag(tag_id: int):
    try:
        storage.delete_tag(tag_id)
        return JSONResponse(status_code=204, content=None)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/tags/template")
async def download_tag_template():
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tags"
        ws.append(["Name", "Description", "Category"])
        ws.append(["Example Tag Name", "What this tag extracts from documents", "Financial"])
        ws.append(["Annual Rent", "The total annual rent amount", "Financial"])
        ws.append(["Lease Start Date", "The date when the lease begins", "Dates"])

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=tag_import_template.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/tags/export")
async def export_tags():
    try:
        from openpyxl import Workbook
        tags_list = storage.get_tags()
        wb = Workbook()
        ws = wb.active
        ws.title = "Tags"
        ws.append(["Name", "Description", "Category"])
        for tag in tags_list:
            ws.append([
                tag.get("name", ""),
                tag.get("description", ""),
                tag.get("category", ""),
            ])

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=tags_export.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tags/upload")
async def upload_tags(file: UploadFile = File(...)):
    try:
        from openpyxl import load_workbook
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        imported = 0
        skipped = 0
        total = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            total += 1
            row_dict = dict(zip(headers, row))
            name = row_dict.get("name") or row_dict.get("Name") or row_dict.get("TAG") or row_dict.get("tag") or row_dict.get("Tag Name") or row_dict.get("Tag")
            if not name:
                vals = [v for v in row if v]
                name = str(vals[0]) if vals else None
            if not name or not isinstance(name, str):
                continue

            description = row_dict.get("description") or row_dict.get("Description") or row_dict.get("desc") or ""
            category = row_dict.get("category") or row_dict.get("Category") or row_dict.get("cat") or ""

            try:
                storage.create_tag(
                    name=str(name).strip(),
                    description=str(description).strip() if description else None,
                    category=str(category).strip() if category else None,
                )
                imported += 1
            except Exception:
                skipped += 1

        return {"imported": imported, "skipped": skipped, "total": total}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/extractions")
async def list_extractions():
    try:
        return storage.get_extractions()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _clean_cell_value(value: str) -> str:
    import re
    if not value:
        return value
    cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', value)
    cleaned = cleaned.replace('\r\n', '\n').replace('\r', '\n')
    cleaned = re.sub(r'[^\x09\x0a\x0d\x20-\x7e\x80-\uffff]', '', cleaned)
    cleaned = cleaned.strip()
    if cleaned and cleaned[0] in ('=', '+', '-', '@'):
        cleaned = "'" + cleaned
    return cleaned


@app.get("/api/extractions/export")
async def export_extractions():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        all_extractions = storage.get_extractions()
        completed = [e for e in all_extractions if e.get("status") == "completed" and e.get("results")]

        if not completed:
            raise HTTPException(status_code=404, detail="No completed extractions to export")

        tag_names = []
        seen_tags = set()
        for ext in completed:
            if ext.get("results"):
                for key in ext["results"]:
                    if key not in seen_tags:
                        seen_tags.add(key)
                        tag_names.append(key)

        wb = Workbook()
        ws = wb.active
        ws.title = "Extractions"

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="D04A02", end_color="D04A02", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        data_font = Font(name="Calibri", size=10)
        data_alignment = Alignment(vertical="top", wrap_text=True)
        alt_fill = PatternFill(start_color="FFF5F0", end_color="FFF5F0", fill_type="solid")
        not_found_font = Font(name="Calibri", size=10, color="999999", italic=True)
        error_font = Font(name="Calibri", size=10, color="CC0000")

        headers = ["Site ID", "Lease Number", "Extraction Date"] + tag_names
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, ext in enumerate(completed, 2):
            ws.cell(row=row_idx, column=1, value=_clean_cell_value(str(ext.get("siteId", "")))).font = data_font
            ws.cell(row=row_idx, column=1).alignment = data_alignment
            ws.cell(row=row_idx, column=1).border = thin_border

            ws.cell(row=row_idx, column=2, value=_clean_cell_value(str(ext.get("leaseNumber", "")))).font = data_font
            ws.cell(row=row_idx, column=2).alignment = data_alignment
            ws.cell(row=row_idx, column=2).border = thin_border

            extracted_at = ext.get("extractedAt", "")
            if extracted_at:
                try:
                    from datetime import datetime as dt
                    parsed = dt.fromisoformat(extracted_at.replace("Z", "+00:00"))
                    extracted_at = parsed.strftime("%Y-%m-%d %H:%M")
                except Exception:
                    pass
            ws.cell(row=row_idx, column=3, value=str(extracted_at)).font = data_font
            ws.cell(row=row_idx, column=3).alignment = data_alignment
            ws.cell(row=row_idx, column=3).border = thin_border

            results = ext.get("results", {}) or {}
            for tag_idx, tag_name in enumerate(tag_names):
                col = tag_idx + 4
                raw_value = results.get(tag_name, "")
                cleaned = _clean_cell_value(str(raw_value)) if raw_value else ""

                cell = ws.cell(row=row_idx, column=col, value=cleaned if cleaned else "Not Found")
                cell.border = thin_border
                cell.alignment = data_alignment

                if not cleaned or cleaned == "Not Found":
                    cell.font = not_found_font
                    cell.value = "Not Found"
                elif "extraction error" in cleaned.lower():
                    cell.font = error_font
                else:
                    cell.font = data_font

            if row_idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = alt_fill

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        col_widths = {"Site ID": 18, "Lease Number": 18, "Extraction Date": 18}
        for col_idx, header in enumerate(headers, 1):
            width = col_widths.get(header, None)
            if width is None:
                max_len = len(header)
                for row_idx in range(2, len(completed) + 2):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val:
                        max_len = max(max_len, min(len(str(val)), 50))
                width = max_len + 4
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = width

        ws.sheet_properties.tabColor = "D04A02"

        if len(completed) > 0:
            summary = wb.create_sheet("Summary")
            summary.sheet_properties.tabColor = "2D2D2D"

            for col_idx, header in enumerate(["Metric", "Value"], 1):
                cell = summary.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
                cell.alignment = header_alignment
                cell.border = thin_border

            summary_data = [
                ("Total Extractions", len(completed)),
                ("Unique Sites", len(set(e.get("siteId", "") for e in completed))),
                ("Total Tags Tracked", len(tag_names)),
            ]

            total_found = 0
            total_missing = 0
            total_errors = 0
            for ext in completed:
                results = ext.get("results", {}) or {}
                for tag_name in tag_names:
                    val = results.get(tag_name, "")
                    v = str(val) if val else ""
                    if "extraction error" in v.lower():
                        total_errors += 1
                    elif v and v != "Not Found" and v.strip():
                        total_found += 1
                    else:
                        total_missing += 1

            total_cells = total_found + total_missing + total_errors
            success_rate = round((total_found / total_cells * 100), 1) if total_cells > 0 else 0

            summary_data.extend([
                ("Tags Found", total_found),
                ("Tags Not Found", total_missing),
                ("Extraction Errors", total_errors),
                ("Success Rate", f"{success_rate}%"),
            ])

            for row_idx, (metric, value) in enumerate(summary_data, 2):
                cell_m = summary.cell(row=row_idx, column=1, value=metric)
                cell_m.font = Font(name="Calibri", size=10, bold=True)
                cell_m.alignment = data_alignment
                cell_m.border = thin_border

                cell_v = summary.cell(row=row_idx, column=2, value=value)
                cell_v.font = data_font
                cell_v.alignment = data_alignment
                cell_v.border = thin_border

                if row_idx % 2 == 0:
                    cell_m.fill = alt_fill
                    cell_v.fill = alt_fill

            summary.column_dimensions["A"].width = 25
            summary.column_dimensions["B"].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=extractions_export.xlsx"}
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/file-type-counts")
async def file_type_counts():
    try:
        return storage.get_file_type_counts_by_lease()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/extractions/{extraction_id}")
async def delete_extraction(extraction_id: int):
    try:
        storage.delete_extraction(extraction_id)
        return JSONResponse(status_code=204, content=None)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/extractions/delete-batch")
async def delete_batch_extractions(request: Request):
    try:
        body = await request.json()
        ids = body.get("ids", [])
        if not ids:
            raise HTTPException(status_code=400, detail="No extraction IDs provided")
        storage.delete_extractions(ids)
        return {"deleted": len(ids)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/files/{file_id}/preview")
async def preview_file(file_id: int):
    try:
        file_record = storage.get_file(file_id)
        if not file_record:
            raise HTTPException(status_code=404, detail="File not found")

        file_path = file_record["file_path"]
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found on disk")

        ext = file_record["file_type"].lower()

        if ext in ("msg", "eml"):
            html = render_email_as_html(file_path, ext)
            return HTMLResponse(content=html)

        mime_types = {
            "pdf": "application/pdf",
            "txt": "text/plain",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        }

        media_type = mime_types.get(ext, "application/octet-stream")
        return FileResponse(
            file_path,
            media_type=media_type,
            headers={
                "Content-Disposition": f'inline; filename="{file_record["file_name"]}"',
                "X-Frame-Options": "SAMEORIGIN",
                "Content-Security-Policy": "frame-ancestors 'self'",
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/files/{file_id}/content")
async def get_file_content(file_id: int):
    try:
        file_record = storage.get_file(file_id)
        if not file_record:
            raise HTTPException(status_code=404, detail="File not found")

        file_path = file_record["file_path"]
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found on disk")

        content = parse_document(file_path, file_record["file_type"])
        return {
            "fileName": file_record["file_name"],
            "fileType": file_record["file_type"],
            "content": content,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/extractions/start/{lease_id}")
async def start_extraction(lease_id: int, request: Request):
    try:
        body = {}
        try:
            body = await request.json()
        except Exception:
            pass
        override_model = body.get("model")
        override_base_url = body.get("baseUrl")

        lease = storage.get_lease(lease_id)
        if not lease:
            raise HTTPException(status_code=404, detail="Lease not found")

        extraction = storage.get_extraction_by_lease(lease_id)
        if extraction and extraction["status"] == "processing":
            if not _is_extraction_stale(extraction):
                raise HTTPException(status_code=409, detail="Extraction already in progress")
            print(f"[EXTRACT] Stale processing extraction detected for lease {lease_id}, restarting")

        if extraction:
            storage.update_extraction(extraction["id"], status="processing", results=None)
        else:
            extraction = storage.create_extraction(lease_id, "processing")

        task_id = generate_task_id()
        ext_id = extraction["id"]
        site_id = lease["site_id"]

        def run_extraction():
            from server_py.config import set_extraction_overrides, clear_extraction_overrides
            acquired = EXTRACTION_SEMAPHORE.acquire(timeout=EXTRACTION_TIMEOUT_MINUTES * 60)
            if not acquired:
                print(f"Extraction timed out waiting for slot: lease {lease_id}")
                emit_progress({
                    "taskId": task_id, "type": "extraction", "status": "failed",
                    "current": 0, "total": 0, "message": "Timed out waiting for extraction slot"
                })
                storage.update_extraction(ext_id, status="failed")
                return
            try:
                if override_model or override_base_url:
                    set_extraction_overrides(model=override_model, base_url=override_base_url)
                results = extract_tags_from_lease(lease_id, task_id, site_id)
                storage.update_extraction(ext_id, status="completed", results=results, extracted_at=datetime.now())
                storage.update_lease_status(lease_id, "extracted")
            except Exception as error:
                print(f"Extraction failed: {error}")
                emit_progress({
                    "taskId": task_id, "type": "extraction", "status": "failed",
                    "current": 0, "total": 0, "message": str(error) or "Extraction failed"
                })
                storage.update_extraction(ext_id, status="failed")
            finally:
                clear_extraction_overrides()
                EXTRACTION_SEMAPHORE.release()

        thread = threading.Thread(target=run_extraction, daemon=True)
        thread.start()

        return {"message": "Extraction started", "extractionId": ext_id, "taskId": task_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/extractions/start-site/{site_id}")
async def start_site_extraction(site_id: int, request: Request):
    try:
        body = {}
        try:
            body = await request.json()
        except Exception:
            pass
        override_model = body.get("model")
        override_base_url = body.get("baseUrl")

        site = storage.get_site(site_id)
        if not site:
            raise HTTPException(status_code=404, detail="Site not found")

        leases_list = storage.get_leases_by_site(site_id)
        if not leases_list:
            raise HTTPException(status_code=400, detail="No leases found for this site")

        task_id = generate_task_id()
        started = []
        skipped = []
        lease_task_ids = {}

        for lease in leases_list:
            extraction = storage.get_extraction_by_lease(lease["id"])
            if extraction and extraction["status"] == "processing":
                if not _is_extraction_stale(extraction):
                    skipped.append(lease["id"])
                    continue
                print(f"[EXTRACT] Stale processing extraction for lease {lease['id']}, restarting")

            if extraction:
                storage.update_extraction(extraction["id"], status="processing", results=None)
            else:
                extraction = storage.create_extraction(lease["id"], "processing")

            started.append(lease["id"])
            lease_task_id = generate_task_id()
            lease_task_ids[str(lease["id"])] = lease_task_id

            ext_id = extraction["id"]
            l_id = lease["id"]

            def run_extraction(eid=ext_id, lid=l_id, tid=lease_task_id, sid=site_id):
                from server_py.config import set_extraction_overrides, clear_extraction_overrides
                acquired = EXTRACTION_SEMAPHORE.acquire(timeout=EXTRACTION_TIMEOUT_MINUTES * 60)
                if not acquired:
                    print(f"Extraction timed out waiting for slot: lease {lid}")
                    emit_progress({
                        "taskId": tid, "type": "extraction", "status": "failed",
                        "current": 0, "total": 0, "message": "Timed out waiting for extraction slot"
                    })
                    storage.update_extraction(eid, status="failed")
                    return
                try:
                    if override_model or override_base_url:
                        set_extraction_overrides(model=override_model, base_url=override_base_url)
                    results = extract_tags_from_lease(lid, tid, sid)
                    storage.update_extraction(eid, status="completed", results=results, extracted_at=datetime.now())
                    storage.update_lease_status(lid, "extracted")
                except Exception as error:
                    print(f"Extraction failed for lease {lid}: {error}")
                    emit_progress({
                        "taskId": tid, "type": "extraction", "status": "failed",
                        "current": 0, "total": 0, "message": str(error) or "Extraction failed"
                    })
                    storage.update_extraction(eid, status="failed")
                finally:
                    clear_extraction_overrides()
                    EXTRACTION_SEMAPHORE.release()

            thread = threading.Thread(target=run_extraction, daemon=True)
            thread.start()

        return {
            "message": "Site extraction started",
            "started": len(started),
            "skipped": len(skipped),
            "taskId": task_id,
            "leaseTaskIds": lease_task_ids,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/extractions/start-sites")
async def start_batch_extraction(request: Request):
    try:
        body = await request.json()
        site_ids = body.get("siteIds", [])
        override_model = body.get("model")
        override_base_url = body.get("baseUrl")
        if not site_ids:
            raise HTTPException(status_code=400, detail="siteIds array is required")

        task_id = generate_task_id()
        total_started = 0
        total_skipped = 0
        lease_task_ids = {}

        for sid in site_ids:
            site = storage.get_site(sid)
            if not site:
                continue

            leases_list = storage.get_leases_by_site(sid)

            for lease in leases_list:
                extraction = storage.get_extraction_by_lease(lease["id"])
                if extraction and extraction["status"] == "processing":
                    if not _is_extraction_stale(extraction):
                        total_skipped += 1
                        continue
                    print(f"[EXTRACT] Stale processing extraction for lease {lease['id']}, restarting")

                if extraction:
                    storage.update_extraction(extraction["id"], status="processing", results=None)
                else:
                    extraction = storage.create_extraction(lease["id"], "processing")

                total_started += 1
                lease_task_id = generate_task_id()
                lease_task_ids[str(lease["id"])] = lease_task_id

                ext_id = extraction["id"]
                l_id = lease["id"]
                current_site_id = sid

                def run_extraction(eid=ext_id, lid=l_id, tid=lease_task_id, csid=current_site_id):
                    from server_py.config import set_extraction_overrides, clear_extraction_overrides
                    acquired = EXTRACTION_SEMAPHORE.acquire(timeout=EXTRACTION_TIMEOUT_MINUTES * 60)
                    if not acquired:
                        print(f"Extraction timed out waiting for slot: lease {lid}")
                        emit_progress({
                            "taskId": tid, "type": "extraction", "status": "failed",
                            "current": 0, "total": 0, "message": "Timed out waiting for extraction slot"
                        })
                        storage.update_extraction(eid, status="failed")
                        return
                    try:
                        if override_model or override_base_url:
                            set_extraction_overrides(model=override_model, base_url=override_base_url)
                        results = extract_tags_from_lease(lid, tid, csid)
                        storage.update_extraction(eid, status="completed", results=results, extracted_at=datetime.now())
                        storage.update_lease_status(lid, "extracted")
                    except Exception as error:
                        print(f"Extraction failed for lease {lid}: {error}")
                        emit_progress({
                            "taskId": tid, "type": "extraction", "status": "failed",
                            "current": 0, "total": 0, "message": str(error) or "Extraction failed"
                        })
                        storage.update_extraction(eid, status="failed")
                    finally:
                        clear_extraction_overrides()
                        EXTRACTION_SEMAPHORE.release()

                thread = threading.Thread(target=run_extraction, daemon=True)
                thread.start()

        return {
            "message": "Batch extraction started",
            "started": total_started,
            "skipped": total_skipped,
            "taskId": task_id,
            "leaseTaskIds": lease_task_ids,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/costs/summary")
async def costs_summary():
    try:
        rows = execute_query("""
            SELECT
                COALESCE(SUM(CASE WHEN type = 'extraction' THEN cost_inr ELSE 0 END), 0)::real AS extraction_inr,
                COALESCE(SUM(CASE WHEN type = 'extraction' THEN cost_usd ELSE 0 END), 0)::real AS extraction_usd,
                COALESCE(SUM(CASE WHEN type = 'extraction' THEN input_tokens ELSE 0 END), 0)::int AS total_input_tokens,
                COALESCE(SUM(CASE WHEN type = 'extraction' THEN output_tokens ELSE 0 END), 0)::int AS total_output_tokens,
                COALESCE(SUM(CASE WHEN type = 'extraction' THEN total_tokens ELSE 0 END), 0)::int AS total_tokens
            FROM cost_logs
        """)
        row = rows[0] if rows else {}
        return {
            "totalInr": row.get("extraction_inr", 0),
            "totalUsd": row.get("extraction_usd", 0),
            "extractionInr": row.get("extraction_inr", 0),
            "extractionUsd": row.get("extraction_usd", 0),
            "totalInputTokens": row.get("total_input_tokens", 0),
            "totalOutputTokens": row.get("total_output_tokens", 0),
            "totalTokens": row.get("total_tokens", 0),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/costs/by-site")
async def costs_by_site():
    try:
        rows = execute_query("""
            SELECT
                s.id AS site_id,
                s.site_id AS site_name,
                COALESCE(SUM(CASE WHEN c.type = 'extraction' THEN c.cost_inr ELSE 0 END), 0)::real AS total_inr,
                COALESCE(SUM(CASE WHEN c.type = 'extraction' THEN c.cost_usd ELSE 0 END), 0)::real AS total_usd,
                COALESCE(SUM(CASE WHEN c.type = 'extraction' THEN c.cost_inr ELSE 0 END), 0)::real AS extraction_inr
            FROM sites s
            LEFT JOIN cost_logs c ON c.site_id = s.id
            GROUP BY s.id, s.site_id
            ORDER BY total_inr DESC
        """)
        return [{
            "siteId": r["site_id"],
            "siteName": r["site_name"],
            "totalInr": r["total_inr"] or 0,
            "totalUsd": r["total_usd"] or 0,
            "extractionInr": r["extraction_inr"] or 0,
        } for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/costs/by-lease/{site_id}")
async def costs_by_lease(site_id: int):
    try:
        rows = execute_query("""
            SELECT
                l.id AS lease_id,
                l.lease_number,
                COALESCE(SUM(CASE WHEN c.type = 'extraction' THEN c.cost_inr ELSE 0 END), 0)::real AS total_inr,
                COALESCE(SUM(CASE WHEN c.type = 'extraction' THEN c.cost_usd ELSE 0 END), 0)::real AS total_usd,
                COALESCE(SUM(CASE WHEN c.type = 'extraction' THEN c.cost_inr ELSE 0 END), 0)::real AS extraction_inr
            FROM leases l
            LEFT JOIN cost_logs c ON c.lease_id = l.id
            WHERE l.site_id = %s
            GROUP BY l.id, l.lease_number
            ORDER BY total_inr DESC
        """, (site_id,))
        return [{
            "leaseId": r["lease_id"],
            "leaseNumber": r["lease_number"],
            "totalInr": r["total_inr"] or 0,
            "totalUsd": r["total_usd"] or 0,
            "extractionInr": r["extraction_inr"] or 0,
        } for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/settings")
async def get_settings():
    try:
        rows = execute_query("SELECT key, value, updated_at FROM app_settings ORDER BY key")
        settings = {}
        for r in rows:
            settings[r["key"]] = r["value"]
        if "usd_to_inr" not in settings:
            settings["usd_to_inr"] = "83.5"
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/settings")
async def update_settings(request: Request):
    try:
        body = await request.json()
        for key, value in body.items():
            execute_no_fetch(
                """INSERT INTO app_settings (key, value, updated_at)
                VALUES (%s, %s, CURRENT_TIMESTAMP)
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP""",
                (key, str(value)),
            )
        from server_py.cost_tracker import invalidate_rate_cache
        invalidate_rate_cache()
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config")
async def get_config_endpoint():
    from server_py.config import get_config_for_api
    try:
        return get_config_for_api()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/config")
async def save_config_endpoint(request: Request):
    from server_py.config import save_config, SENSITIVE_KEYS, get_config
    try:
        body = await request.json()
        current = get_config()
        for key in SENSITIVE_KEYS:
            if key in body:
                val = body[key]
                if val and val.startswith("*"):
                    body[key] = current.get(key, "")
                elif val == "set" or val == "":
                    body[key] = current.get(key, "")
        save_config(body)
        from server_py.cost_tracker import invalidate_rate_cache
        invalidate_rate_cache()
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/config/reset-prompts")
async def reset_prompts_endpoint():
    from server_py.config import reset_prompts_to_default
    try:
        reset_prompts_to_default()
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/settings/recalculate-costs")
async def recalculate_costs():
    from server_py.cost_tracker import get_usd_to_inr
    try:
        rate = get_usd_to_inr()
        execute_no_fetch(
            "UPDATE cost_logs SET cost_inr = cost_usd * %s",
            (rate,),
        )
        return {"success": True, "rate": rate}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/progress/{task_id}")
async def progress_sse(task_id: str):
    async def event_generator():
        queue = asyncio.Queue()
        loop = asyncio.get_event_loop()

        def callback(data):
            try:
                loop.call_soon_threadsafe(queue.put_nowait, data)
            except Exception:
                pass

        subscribe(task_id, callback)
        try:
            while True:
                try:
                    data = await asyncio.wait_for(queue.get(), timeout=30)
                    yield f"data: {json.dumps(data)}\n\n"
                    if data.get("status") in ("completed", "failed"):
                        break
                except asyncio.TimeoutError:
                    yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"
        finally:
            unsubscribe(task_id, callback)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


def _map_extraction_response(ext):
    if not ext:
        return None
    return {
        "id": ext["id"],
        "leaseId": ext["lease_id"],
        "status": ext["status"],
        "results": ext.get("results"),
        "extractedAt": ext["extracted_at"].isoformat() if ext.get("extracted_at") else None,
        "createdAt": ext["created_at"].isoformat() if ext.get("created_at") else None,
    }


DIST_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "dist", "public")

if os.environ.get("NODE_ENV") == "production" and os.path.exists(DIST_PATH):
    app.mount("/assets", StaticFiles(directory=os.path.join(DIST_PATH, "assets")), name="assets")

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        file_path = os.path.join(DIST_PATH, full_path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(DIST_PATH, "index.html"))


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", "5000"))
    uvicorn.run(app, host="0.0.0.0", port=port)
